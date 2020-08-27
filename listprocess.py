from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill



lightblueFill = PatternFill(start_color='50bcdf', end_color='50bcdf', fill_type='solid')


load_wb = load_workbook("fileToRead.xlsx", data_only=True)
write_wb = Workbook()
load_ws = load_wb['Sheet1']
write_ws = write_wb.active

nomatchlist = ['데상트 여성트레이닝복세트', '르꼬끄 여성트레이닝복세트', '휠라 여성트레이닝복세트', '시크릿화장품 소금', '시크릿화장품 비누', '시크릿화장품 쿠션', '시크릿화장품 수분크림', '풀햄 히트텍', '헤비추얼 남성데님', '헤비추얼 툴라이트', '샤넬 호보백', '조이풀샌들', 
'조이풀슬립온', '이디야 커피쿠폰', '삼성에어컨 덮개', '삼성에어컨 리모컨', '키스포 기모', '키스포 기모세트', '키스포 기모바지', '키스포 기모티', '키스포 기모후드', '키스포 기모바지5종', '키스포 기모맨투맨', '밀가루 20kg', '유기농 밀가루', '밀가루 보관용기', 
'벨라지오 각티슈', '벨라지오 화장지']


get_cells = load_ws['B364':'H400']
for r in get_cells:
    for cell in r:
        val = cell.value
        #keywords.append(val)
        write_ws.cell(row = cell.row, column = cell.column).value = val


get_cells = load_ws['B364':'H400']
for r in get_cells:
    for cell in r:
        val = cell.value
        if val in nomatchlist:
            #write_ws.cell(row = cell.row, column = cell.column).value = val
            write_ws.cell(row = cell.row, column = cell.column).fill = lightblueFill



write_wb.save("fileToWrite.xlsx")
