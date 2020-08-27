import urllib.parse
from bs4 import BeautifulSoup
from urllib.request import urlopen

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill

lightblueFill = PatternFill(start_color='50bcdf', end_color='50bcdf', fill_type='solid')


load_wb = load_workbook("fileToRead.xlsx", data_only=True)
write_wb = Workbook()
load_ws = load_wb['Sheet1']
write_ws = write_wb.active

keywords = []

get_cells = load_ws['B601':'H684']
for r in get_cells:
    for cell in r:
        val = cell.value
        keywords.append(val)
        write_ws.cell(row = cell.row, column = cell.column).value = val
            
#write_wb.save("fileToWrite.xlsx")


# 엑셀에서 불러온 키워드를 keyword에 저장
#print(keywords)
notmatchkeywordstr = ""
notmatchkeyword = []
f = open("nomatch.txt", 'w')
    
for keyword in keywords:
    flag = "yes"
    basic_url = "http://m.hnsmall.com/search?query_top="
    urlToOpen = basic_url + keyword
    link = urllib.parse.quote(urlToOpen, safe=':/?-=')

    with urlopen(link) as response:
        soup = BeautifulSoup(response, 'html.parser')
        i = 1
        #total_data = " "
        filetowrite = keyword + ".txt"
        
        #print(soup.select('p.title'))
        print("key word is {}".format(keyword))
        
        if soup.select('p.title') == []:
            #print(keyword)
            notmatchkeyword.append(keyword)
            print(notmatchkeyword)
            notmatchkeywordstr = "".join(notmatchkeyword)
            notmatchkeywordstr += "\n\n"
    f.write(notmatchkeywordstr)
print(notmatchkeyword)
f.close()

get_cells = load_ws['B601':'H684']
for r in get_cells:
    for cell in r:
        val = cell.value
        if val in notmatchkeyword:
            #write_ws.cell(row = cell.row, column = cell.column).value = val
            write_ws.cell(row = cell.row, column = cell.column).fill = lightblueFill



write_wb.save("fileToWrite.xlsx")

    
    